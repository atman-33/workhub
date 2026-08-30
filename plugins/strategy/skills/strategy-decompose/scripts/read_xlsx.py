#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Read a decomposition workbook back into the intermediate JSON.

Usage:
    python read_xlsx.py <input.xlsx> <output.json> [--sheet NAME]

Use this to pick up edits someone made in Excel before auditing or rebuilding.
The header row is located by content, not by position: a workbook a human has
edited often carries a note on row 1, a blank row 2 and the header on row 3.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: python -m pip install openpyxl")

HEADER_HINTS = ("level 1", "level 2", "strategy", "domain")


def find_header_row(ws, limit=12):
    """Return the 1-based row index of the header, or None."""
    for row in range(1, min(ws.max_row, limit) + 1):
        cells = [str(ws.cell(row, col).value or "").lower() for col in range(1, 6)]
        joined = " ".join(cells)
        if any(hint in joined for hint in HEADER_HINTS) and "level" in joined:
            return row
    return None


def read(ws):
    header = find_header_row(ws)
    if header is None:
        sys.exit("could not find the header row (looked for a row mentioning 'Level')")
    note = ws.cell(1, 1).value if header > 1 else ""

    data = {"sheet": ws.title, "header_note": note or "", "strategies": []}
    for row in range(header + 1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, 10)]
        if not any(values):
            continue
        level1, level2, level3, level4, level5, memo, priority, rationale, note_col = values
        if level1:
            data["strategies"].append({"level1": level1, "domains": []})
        if not data["strategies"]:
            sys.exit(f"row {row} has content before any Level 1 value")
        strategy = data["strategies"][-1]
        if level2:
            strategy["domains"].append({"level2": level2, "components": []})
        domain = strategy["domains"][-1]
        if level3:
            domain["components"].append({"level3": level3, "aspects": []})
        component = domain["components"][-1]
        if level4:
            component["aspects"].append({"level4": level4, "states": []})
        aspect = component["aspects"][-1]
        if level5:
            aspect["states"].append({
                "level5": level5,
                "memo": memo or "",
                "priority": priority or "",
                "rationale": rationale or "",
                "note": note_col or "",
            })
    return data


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="the workbook to read")
    parser.add_argument("output", help="where to write the intermediate JSON")
    parser.add_argument("--sheet", default=None, help="sheet name (default: the active sheet)")
    args = parser.parse_args()

    wb = load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active
    data = read(ws)
    Path(args.output).write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")

    states = sum(
        len(aspect["states"])
        for strategy in data["strategies"]
        for domain in strategy["domains"]
        for component in domain["components"]
        for aspect in component["aspects"]
    )
    print(f"read {len(data['strategies'])} strategies / {states} states from {args.input} (sheet: {ws.title})")


if __name__ == "__main__":
    main()
