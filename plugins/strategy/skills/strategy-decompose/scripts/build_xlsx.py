#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a strategy decomposition workbook from the intermediate JSON.

Usage:
    python build_xlsx.py <input.json> <output.xlsx> [--sheet NAME]

The JSON schema is documented in ../references/xlsx-format.md.
Rows repeat a parent value only when it changes, so the reader can see
where each branch starts.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: python -m pip install openpyxl")

HEADERS = [
    "Strategy / phase (Level 1)",
    "Domain (Level 2)",
    "Component (Level 3)",
    "Aspect (Level 4)",
    "Target state (Level 5)",
    "Memo",
    "Priority",
    "Rationale (upper connection / local situation)",
    "Notes",
]
WIDTHS = [34, 32, 38, 42, 78, 24, 10, 60, 30]
FILLS = {
    "blue": "E8F1FB",
    "orange": "FDF0E6",
    "green": "E8F5E9",
    "purple": "F3E8FB",
    "gray": "EFEFEF",
}
FILL_ORDER = ["blue", "orange", "green", "purple", "gray"]


def flatten(data):
    """Yield one row (list of 9 values) per Level 5 state."""
    rows = []
    for index, strategy in enumerate(data.get("strategies", [])):
        color = strategy.get("color") or FILL_ORDER[index % len(FILL_ORDER)]
        first_of_strategy = True
        for domain in strategy.get("domains", []):
            first_of_domain = True
            for component in domain.get("components", []):
                first_of_component = True
                for aspect in component.get("aspects", []):
                    first_of_aspect = True
                    for state in aspect.get("states", []):
                        rows.append((
                            color,
                            [
                                strategy.get("level1") if first_of_strategy else None,
                                domain.get("level2") if first_of_domain else None,
                                component.get("level3") if first_of_component else None,
                                aspect.get("level4") if first_of_aspect else None,
                                state.get("level5"),
                                state.get("memo"),
                                state.get("priority"),
                                state.get("rationale"),
                                state.get("note"),
                            ],
                        ))
                        first_of_strategy = False
                        first_of_domain = False
                        first_of_component = False
                        first_of_aspect = False
    return rows


def build(data, out_path, sheet_name):
    rows = flatten(data)
    if not rows:
        sys.exit("no Level 5 states found in the input JSON")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    note = data.get("header_note") or ""
    ws.cell(1, 1, note)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.cell(1, 1).font = Font(bold=True, size=11)
    ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(2, col, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="44546A")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    ws.row_dimensions[2].height = 30

    for offset, (color, values) in enumerate(rows):
        row = offset + 3
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col)
            if value not in (None, ""):
                cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=FILLS.get(color, FILLS["gray"]))
        for col in (1, 2):
            if ws.cell(row, col).value:
                ws.cell(row, col).font = Font(bold=True)

    for col, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "B3"

    wb.save(out_path)
    return len(rows)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="intermediate JSON produced in phase 3")
    parser.add_argument("output", help="target .xlsx (write to the AI-owned -draft file)")
    parser.add_argument("--sheet", default=None, help="sheet name (default: from JSON, else Decomposition)")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    sheet = args.sheet or data.get("sheet") or "Decomposition"
    count = build(data, args.output, sheet)
    print(f"wrote {count} rows to {args.output} (sheet: {sheet})")


if __name__ == "__main__":
    main()
